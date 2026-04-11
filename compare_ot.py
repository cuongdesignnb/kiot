import openpyxl, re

wb = openpyxl.load_workbook(r'd:\Kiot\kiotviet-clone\BangChiTietChamCong_KV11042026-074801-768.xlsx')
ws = wb['Bảng chi tiết chấm công']

def parse_ot(s):
    if not s or s == '-': return 0
    h = m = 0
    hm = re.match(r'(\d+)h(\d+)p', str(s))
    mm = re.match(r'(\d+)p', str(s))
    hh = re.match(r'(\d+)h', str(s))
    if hm: h,m = int(hm.group(1)), int(hm.group(2))
    elif mm: m = int(mm.group(1))
    elif hh: h = int(hh.group(1))
    return h*60+m

# Our debug-ot data for NXT (before, after, total)
our_nxt = {
    '01': (5,0,5), '02': (2,75,77), '03': (7,0,7), '04': (6,41,47), '05': (3,0,3),
    '10': (8,83,91), '11': (7,51,58), '12': (2,38,40), '13': (6,32,38), '14': (5,0,5),
    '15': (0,0,0), '16': (7,68,75), '17': (2,117,119), '18': (0,78,78), '19': (8,110,118),
    '20': (9,94,103), '21': (7,42,49), '22': (4,0,4), '23': (6,29,35), '24': (8,53,61),
    '25': (4,63,67), '26': (3,102,105), '27': (0,0,0), '28': (6,26,32), '29': (2,12,14),
    '30': (2,0,2), '31': (4,38,42),
}

our_tuan = {
    '02': (1,0,1), '03': (0,30,30), '04': (0,9,9), '05': (0,0,0),
    '07': (0,95,95), '10': (4,84,88), '11': (0,51,51), '12': (0,39,39),
    '13': (0,10,10), '14': (0,0,0), '16': (0,67,67), '17': (0,14,14),
    '20': (3,0,3), '21': (0,25,25), '24': (5,0,5), '25': (0,11,11),
    '28': (0,26,26), '31': (0,11,11),
}

# Employee sections: NV000024 rows 4-35, NV000028 rows 36-67, NV000026 rows 68-99
sections = {
    'NV000026': (69, 99, our_nxt, 'NGUYEN XUAN THANH'),
    'NV000028': (37, 67, our_tuan, 'TRAN ANH TUAN'),
}

for code, (start_row, end_row, our_data, name) in sections.items():
    print(f"\n{'='*90}")
    print(f"=== {name} ({code}) ===")
    print(f"{'='*90}")
    print(f"{'Date':>5} | {'DOW':>10} | {'KV_OT':>6} | {'US_OT':>6} | {'BEF':>3} | {'AFT':>4} | {'DIFF':>5} | Note")
    print("-" * 90)
    
    total_kv = 0
    total_us = 0
    
    for r in range(start_row, end_row + 1):
        date_raw = ws.cell(r, 5).value
        if not date_raw: continue
        
        day_val = ws.cell(r, 6).value or ''
        ot_str = ws.cell(r, 16).value or '-'
        kv_ot = parse_ot(str(ot_str))
        
        dd = str(date_raw)[:2]
        bef, aft, us_ot = our_data.get(dd, (0,0,0))
        diff = us_ot - kv_ot
        total_kv += kv_ot
        total_us += us_ot
        
        note = ''
        if diff > 0:
            note = f'WE +{diff} (bef={bef})'
        elif diff < 0:
            note = f'WE {diff}'
        elif kv_ot > 0 and bef > 0:
            note = f'OK (bef={bef})'
        
        marker = '***' if diff != 0 else '   '
        print(f"{dd}/03 | {day_val:>10} | {kv_ot:>5}m | {us_ot:>5}m | {bef:>3} | {aft:>4} | {diff:>+5} | {marker} {note}")
    
    print("-" * 90)
    diff_total = total_us - total_kv
    print(f"TOTAL |            | {total_kv:>5}m | {total_us:>5}m |     |      | {diff_total:>+5} |")
    print(f"KV: {total_kv}min = {total_kv//60}h{total_kv%60}p | US: {total_us}min = {total_us//60}h{total_us%60}p | DIFF: {diff_total}min")
